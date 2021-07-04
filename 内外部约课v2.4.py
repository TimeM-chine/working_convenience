'''
约课 v2.4 更新内容：
    考虑到【内部约课】需求大幅降低，本版【取消】了内部约课的选项，但将其作为辅助功能编写在程序中。
    本程序能够检测到正式约课的异常如下：
        1. 课时券不足； 2.预约已存在； 3.预约时间异常（如18:00); 4.无课程线； 5.时段已达上限； 6.未导入排班
    对于4、5、6情况的学生，正式约课失败后，会自动进入内部约课，想要关闭此功能可将第14行 auto_in 设置为False。
    以上情况都会写入 内部约课日志.txt，请及时查看。

若cookie过期可以在这个网址寻找，推荐找auth里的cookie：https://maolaozu.codemao.cn/book_appointment/add_edit
'''

# cookie 在这里修改
cookie = '_ga=GA1.2.2057943407.1595661913; SL_C_23361dd035530_KEY=be556a167e74fcde3a3444e29b25f8e99fb0c59f; __guid=111581274.4233868734248029700.1610877138168.5208; gr_user_id=0468cfb0-8575-4ff9-9820-e56a8c44cf0e; Hm_lvt_5159ede36d7853549fcc4db3b12c530c=1623390280; __ca_uid_key__=f18ae27f-12ae-434f-a2f6-467195a7f0be; _gid=GA1.2.138291067.1625047191; internal_account_token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJUb2tlbiIsImF1dGgiOiJST0xFX0FETUlOIiwibmFtZSI6IumCueaxleaWjCIsImVuaWQiOjQ4NDgsImlhdCI6MTYyNTM2MjcxNywianRpIjoiNmYwZGFhMDYtODhhZS00OWMxLWE0ZmUtZDA5MzI0ZjQ2MGQyIn0.A7eO1G539ThePk4EJt16jZ1PI7mH4imMPRPuHr6Pkes; admin-authorization=Bearer+eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJ1c2VyX3R5cGUiOiJhZG1pbiIsInVzZXJfaWQiOjQyODgsImlhdCI6MTYyNTMzMzkxNywianRpIjoiOGRjYTMxOTgtZGM2OC0xMWViLWFjNTUtOTUyMTkzNGM3OWQ2In0.qx8YHvXAzLBpb5fpjL04RcFkfKvCiQgE9EK89wch1Yc; acw_tc=2f624a1816253897791503802e267ea689b9dbfbf5b06f8eea07ebc0b0a357'
auto_in = True  # 设置为True：正式约课失败会自动进行内部约课；设置为False：正式约课失败无任何操作。
yk_n = 4 # 约课次数，若按周几约课，默认预约4次，相当于1个月

import requests
import json
import time
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
# import pprint

# 学生信息准备
wb = load_workbook(r'内部约课表格.xlsx', read_only=True)
sheetnames = wb.sheetnames[0]
ws = wb.get_sheet_by_name(sheetnames)
rows = ws.max_row
columns = ws.max_column

id_list = []
plan_time = []  # 几点
plan_day = []   # 哪天
plan_week_day = []  # 周几
teacher_id = 0
for i in range(2, rows+1):  # 从第二行开始获取数据
    id_list.append(ws.cell(i, 1).value)
    plan_time.append(ws.cell(i, 4).value)
    plan_day.append(ws.cell(i, 5).value)
    plan_week_day.append(ws.cell(i, 6).value)
teacher_id = ws.cell(2, 7).value
# print(id_list,plan_time,plan_day,teacher_id)
if teacher_id == None:
    print("老师id都没填，搁这闹呢？重新运行吧。")
    sys.exit()

# 准备部分
headers = {
    "Accept": "application/json, text/plain, */*",
    "authorization_type": "3",
    "cookie": cookie,
    "content-type": "application/json;charset=UTF-8",
    "Origin": "https://crm.codemao.cn",
    "Referer": "https://crm.codemao.cn/",
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36',
}


# first step 获取课程线id
def get_sub_id(stu_id):
    url_subjects = f'https://api-education-codemaster.codemao.cn/admin/lessons/student-position/users/{stu_id}/subjects'
    r = requests.get(url_subjects, headers=headers)
    if r.reason == 'Unauthorized':
        print("未认证成功，请更新cookie后再试。")
        sys.exit()
    return r.json()['subjects'][0]['id']


# second step 获取课程系列id
def get_track_id(stu_id,sub_id):
    url_tracks = f'https://api-education-codemaster.codemao.cn/admin/lessons/student-position/users/{stu_id}/tracks?subject_id={sub_id}'
    r = requests.get(url_tracks, headers=headers)
    return (r.json()['tracks'][0]['id'])

# third step 获取课程id以及名称
# def get_points(track_id):
#     url_points = f'https://api-education-codemaster.codemao.cn/admin/lessons/student-position/users/{stu_id}/points?track_id={track_id}'
#     r = requests.get(url_points, headers=headers)
#     res = r.json()['points']
#     point_id = res[0]['id']
#     return point_id


def num2time(num):  # 时间戳转日期
    lt = time.localtime(num)
    last_time = time.strftime("%Y-%m-%d %H:%M", lt)
    return last_time


t = time.time()
onehour = 3600
halfhour = 1800
oneday = onehour*24
oneweek = oneday*7

#fourth step 计算约课时间戳
def cal_timeslot(time, day):
    next_mon = 1612108800  # 周一初始日期
    day = day-1
    next_mon += day*oneday
    while next_mon < t-oneday:
        next_mon += oneweek  # 计算下一个周几的时间戳
    if time == 17 or time == 19:
        time += 0.5
    next_mon += time*onehour  # 第一次约课的时间戳
    yuyue_list = []
    for i in range(yk_n):  # 这里设置约课次数
        yuyue_list.append(next_mon)
        next_mon += oneweek
    return yuyue_list


def date2timeslot(clock, date):
    time_array = time.strptime(str(date), "%Y-%m-%d %H:%M:%S")
    timestamp = int(time.mktime(time_array))
    if clock == 17 or clock == 19:
        clock += 0.5
    timestamp += clock*onehour
    a = []
    a.append(timestamp)  # 为了对应下面的程序，这里将时间戳放进列表中
    return a


with open('内部约课日志.txt', 'a') as f:
    f.write("===========以上为历史约课结果============\n")
# 开始约课

url_out = 'https://api-attendance-codemaster.codemao.cn/attendances'
url_in = 'https://api-attendance-codemaster.codemao.cn/attendances/inside'
fail_id_list = []
fail_plan_time = []  # 几点
fail_plan_day = []   # 哪天
fail_plan_week_day = []  # 周几
def change(f_id,f_plan_time,f_plan_day,f_plan_week_day):
    fail_id_list.append(f_id)
    fail_plan_time.append(f_plan_time)
    fail_plan_day.append(f_plan_day)
    fail_plan_week_day.append(f_plan_week_day)
def submit(url_2,id_list,plan_time,plan_day,plan_week_day):
    count = 0
    for i in range(len(id_list)):
        mes = ''
        stu_id = id_list[i]
        if not isinstance(stu_id,int):
            if stu_id == None:
                print("excel表的第",i+2,"行有空值，下次运行请删除，本次跳过")
                continue
            else:
                print("excel表的第",i+2,"行数据有误，请检查，本次跳过")
                continue
        sub_id = get_sub_id(stu_id)
        track_id = get_track_id(stu_id,sub_id)
        if plan_day[i] != None:  # 如果 约课日期 不为空
            yuyue_list = date2timeslot(plan_time[i], plan_day[i])
        elif plan_week_day[i] != None:  # 如果 周几不为空
            yuyue_list = cal_timeslot(plan_time[i], plan_week_day[i])
        else:
            print("学生 " + str(id_list[i])+" 未填约课时间，请检查表格，本次跳过")
            continue
        for j in yuyue_list:
            data = {
                # 'point_id': '2688',  # 课程id  2688表示汉堡这节课   1549表示绘制小车这节课
                'student_id': str(id_list[i]),
                'subject_id': str(sub_id),  # 课程线id  34表示G6
                'teacher_id': str(teacher_id),
                'time_slot': int(j),  # 约课时间戳
                'track_id': str(track_id)   # 课程系列id  46表示Python课G6-1
            }
            # print(data)
            r = requests.post(url_2, headers=headers,data=json.dumps(data))  # 正常无返回值，不能json()
            if r.reason == 'Unauthorized':
                print("未认证成功，请更新cookie后再试。")
                sys.exit()
            # print(r.json()['error_code'])
            if r.status_code == 422:
                # print(r.json())
                if r.json()['error_code'] == r'Not-Enough-Ticket@Attendance-Service':
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  预约时课时券不足\n'
                elif r.json()['error_code'] == "Attendance-Already-Existed@Attendance-Service":
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  预约已存在，请检查约课状态\n'
                elif r.json()['error_code'] == "TimeSlot-Invalid@Attendance-Service":
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  预约时间非正常，请检查时间\n'
                elif r.json()['error_code'] == "Teacher-Not-Belong-Point@Attendance-Service":
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  老师无此课程线，请检查\n'
                    change(id_list[i],plan_time[i],plan_day[i],plan_week_day[i])
                elif r.json()['error_code'] == "Teacher-Overload@Attendance-Service":
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  老师此时段已达上限，请检查\n'
                    change(id_list[i],plan_time[i],plan_day[i],plan_week_day[i])
                elif r.json()['error_code'] == "Teacher-Not-Schedule@Attendance-Service":
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  老师当天还未导入排班，无法约课\n'
                    change(id_list[i],plan_time[i],plan_day[i],plan_week_day[i])
                with open('内部约课日志.txt', 'a') as f:
                    f.write(mes)
            elif r.status_code == 204:
                pass  # 预检请求，不会影响结果
            else:
                print("出现了预料之外的情况，请联系管理员")
                with open('内部约课日志.txt', 'a') as f:
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  未知错误，请联系管理员\n'
                    f.write(mes)
                sys.exit()
        count += 1
        print(count,mes[:-1])

submit(url_out,id_list,plan_time,plan_day,plan_week_day)
if auto_in:
    print('正式约课已结束，发现',str(len(fail_id_list)),'条数据约课失败,即将进行内部约课...')
    with open('内部约课日志.txt', 'a') as f:
        f.write('         ~~~~~~本次内部约课日志如下~~~~~~~   \n')
    submit(url_in, fail_id_list, fail_plan_time,fail_plan_day, fail_plan_week_day)
