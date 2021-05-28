# cookie 在这里修改
cookie = ''  # 这一行可以用来复制，方便填写
cookie = ''

yk_n = 4 # 约课次数，若按周几约课，默认预约4次

import requests
import json
import time
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
# import pprint

# 学生信息准备
wb = load_workbook(r'内部约课表格.xlsx', read_only=True)
sheetnames = wb.sheetnames[0]
ws = wb.get_sheet_by_name(sheetnames)
rows = ws.max_row
columns = ws.max_column

id_list = []
plan_time = []  # 几点
plan_day = []   # 哪天
plan_week_day = []  # 周几
teacher_id = 0
for i in range(2, rows+1):  # 从第二行开始获取数据
    id_list.append(ws.cell(i, 1).value)
    plan_time.append(ws.cell(i, 4).value)
    plan_day.append(ws.cell(i, 5).value)
    plan_week_day.append(ws.cell(i, 6).value)
teacher_id = ws.cell(2, 7).value
# print(id_list,plan_time,plan_day,teacher_id)
if teacher_id == None:
    print("老师id都没填，搁这闹呢？")
    sys.exit()

# 准备部分
headers = {
    "Accept": "application/json, text/plain, */*",
    "authorization_type": "3",
    "cookie": cookie,
    "content-type": "application/json;charset=UTF-8",
    "Origin": "https://crm.codemao.cn",
    "Referer": "https://crm.codemao.cn/",
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36',
}


# first step 获取课程线id
def get_sub_id(stu_id):
    url_subjects = f'https://api-education-codemaster.codemao.cn/admin/lessons/student-position/users/{stu_id}/subjects'
    r = requests.get(url_subjects, headers=headers)
    if r.reason == 'Unauthorized':
        print("未认证成功，请更新cookie后再试。")
        sys.exit()
    return r.json()['subjects'][0]['id']


# second step 获取课程系列id
def get_track_id(stu_id,sub_id):
    url_tracks = f'https://api-education-codemaster.codemao.cn/admin/lessons/student-position/users/{stu_id}/tracks?subject_id={sub_id}'
    r = requests.get(url_tracks, headers=headers)
    return (r.json()['tracks'][0]['id'])

# third step 获取课程id以及名称
# def get_points(track_id):
#     url_points = f'https://api-education-codemaster.codemao.cn/admin/lessons/student-position/users/{stu_id}/points?track_id={track_id}'
#     r = requests.get(url_points, headers=headers)
#     res = r.json()['points']
#     point_id = res[0]['id']
#     return point_id


def num2time(num):  # 时间戳转日期
    lt = time.localtime(num)
    last_time = time.strftime("%Y-%m-%d %H:%M", lt)
    return last_time


t = time.time()
onehour = 3600
halfhour = 1800
oneday = onehour*24
oneweek = oneday*7

#fourth step 计算约课时间戳
def cal_timeslot(time, day):
    next_mon = 1612108800  # 周一初始日期
    day = day-1
    next_mon += day*oneday
    while next_mon < t-oneday:
        next_mon += oneweek  # 计算下一个周几的时间戳
    if time == 17 or time == 19:
        time += 0.5
    next_mon += time*onehour  # 第一次约课的时间戳
    yuyue_list = []
    for i in range(yk_n):  # 这里设置约课次数
        yuyue_list.append(next_mon)
        next_mon += oneweek
    return yuyue_list


def date2timeslot(clock, date):
    time_array = time.strptime(str(date), "%Y-%m-%d %H:%M:%S")
    timestamp = int(time.mktime(time_array))
    if clock == 17 or clock == 19:
        clock += 0.5
    timestamp += clock*onehour
    a = []
    a.append(timestamp)  # 为了对应下面的程序，这里将时间戳放进列表中
    return a


with open('内部约课日志.txt', 'a') as f:
    f.write("===========以上为历史约课结果============\n")
# 开始约课
count = 0
for i in range(len(id_list)):
    stu_id = id_list[i]
    if not isinstance(stu_id,int):
        if stu_id == None:
            print("excel表的第",i+2,"行有空值，下次运行请删除，本次跳过")
            continue
        else:
            print("excel表的第",i+2,"行数据有误，请检查，本次跳过")
            continue
    sub_id = get_sub_id(stu_id)
    track_id = get_track_id(stu_id,sub_id)
    if plan_day[i] != None:  # 如果 约课日期 不为空
        yuyue_list = date2timeslot(plan_time[i], plan_day[i])
    elif plan_week_day[i] != None:  # 如果 周几不为空
        yuyue_list = cal_timeslot(plan_time[i], plan_week_day[i])
    else:
        print("学生 " + str(id_list[i])+" 未填约课时间，请检查表格，本次跳过")
        continue
    for j in yuyue_list:
        data = {
            # 'point_id': '2688',  # 课程id  2688表示汉堡这节课   1549表示绘制小车这节课
            'student_id': str(id_list[i]),
            'subject_id': str(sub_id),  # 课程线id  34表示G6
            'teacher_id': str(teacher_id),
            'time_slot': int(j),  # 约课时间戳
            'track_id': str(track_id)   # 课程系列id  46表示Python课G6-1
        }
        # print(data)
        # 约课链接
        url_2 = 'https://api-attendance-codemaster.codemao.cn/attendances/inside'
        r = requests.post(url_2, headers=headers,data=json.dumps(data))  # 正常无返回值，不能json()
        if r.reason == 'Unauthorized':
            print("未认证成功，请更新cookie后再试。")
            sys.exit()
        if r.status_code == 422:
            # print(r.json())
            if r.json()['error_code'] == r'Not-Enough-Ticket@Attendance-Service':
                with open('内部约课日志.txt', 'a') as f:
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  预约时课时券不足\n'
                    f.write(mes)
            if r.json()['error_code'] == "Attendance-Already-Existed@Attendance-Service":
                with open('内部约课日志.txt', 'a') as f:
                    mes = str(id_list[i])+"  " + \
                        str(num2time(int(j)))+'  预约已存在，请检查约课状态\n'
                    f.write(mes)
        elif r.status_code == 204:
            pass
        else:
            print("出现了预料之外的情况，请联系管理员")
            with open('内部约课日志.txt', 'a') as f:
                mes = str(id_list[i])+"  " + \
                    str(num2time(int(j)))+'  未知错误，请联系管理员\n'
                f.write(mes)
            sys.exit()
    count += 1
    print(count)
